#!/usr/bin/env python3
"""Convert the old workbook-style grid into normalized JSON.

This script is intentionally conservative. It treats placeholders such as XX,
XXX, NA, and -- as non-available values rather than numbers.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

PLACEHOLDERS = {"XX", "XXX", "NA", "N/A", "--", "--%", ""}

STATE_NAMES = {
    "GA": "Georgia",
    "NC": "North Carolina",
    "PA": "Pennsylvania",
    "MI": "Michigan",
    "OH": "Ohio",
    "WI": "Wisconsin",
    "FL": "Florida",
    "AZ": "Arizona",
    "NV": "Nevada",
    "National": "National",
}

METRIC_MAP = {
    "% of Electorate": ("electorate_share", None),
    "DEM": ("vote_share", "Dem"),
    "GOP": ("vote_share", "GOP"),
}


def normalize_value(value: Any) -> tuple[float | None, str]:
    if value is None:
        return None, "missing"
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in PLACEHOLDERS:
            return None, "not_available"
        try:
            value = float(stripped.replace("%", ""))
        except ValueError:
            return None, "not_available"
    if isinstance(value, (int, float)):
        # Workbook percentages are usually stored as decimals. Convert to pct points.
        number = float(value)
        if 0 <= number <= 1:
            number *= 100
        return round(number, 1), "ok"
    return None, "not_available"


def extract_dataset(path: Path, sheet_name: str, value_col: int, label: str, status: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    rows: list[dict[str, Any]] = []

    for block_start_col in (1, 7):
        category_col = block_start_col
        group_col = block_start_col + 1
        data_col = value_col if block_start_col == 1 else value_col + 6
        current_category = None
        for r in range(3, ws.max_row + 1):
            category_value = ws.cell(r, category_col).value
            if category_value not in (None, " ", ""):
                current_category = str(category_value).strip()
            group_value = ws.cell(r, group_col).value
            if group_value is None or current_category is None:
                continue
            group_text = str(group_value).strip()
            if group_text not in METRIC_MAP:
                continue
            metric, party = METRIC_MAP[group_text]
            value, value_status = normalize_value(ws.cell(r, data_col).value)
            rows.append({
                "category": "Workbook group",
                "group": current_category.title(),
                "metric": metric,
                "party": party,
                "value": value,
                "unit": "pct",
                "status": value_status if value_status != "ok" else status,
                "notes": "Imported from legacy workbook layout. Verify before locking."
            })

    return {
        "metadata": {
            "dataset_id": f"{label.lower().replace(' ', '-')}-{sheet_name.lower()}",
            "election_id": label,
            "state": sheet_name,
            "state_name": STATE_NAMES.get(sheet_name, sheet_name),
            "source": "Legacy workbook import",
            "source_url": None,
            "data_status": status,
            "last_updated": datetime.now(timezone.utc).isoformat(),
            "schema_version": "0.1.0"
        },
        "rows": rows
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default="GA")
    parser.add_argument("--value-col", type=int, default=3, help="1-indexed column in the left data block, e.g. 3 for 2024, 5 for 2020")
    parser.add_argument("--label", required=True)
    parser.add_argument("--status", default="draft")
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    dataset = extract_dataset(args.workbook, args.sheet, args.value_col, args.label, args.status)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(dataset, indent=2) + "\n")


if __name__ == "__main__":
    main()
